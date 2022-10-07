import openpyxl
# pip install openpyxl

def main():
    wb = openpyxl.load_workbook("C:\\Users\\kurs\\Documents\\temp\\dane1.xlsx")
    #print(wb.sheetnames)
    #ws = wb.active
    ws = wb["Arkusz3"]
    print(ws['A2'].value)  # tu moga byc obiekty datetime.datetime
    print(ws.cell(row=2, column=1).value)  # indeksowanie od 1 (!)
    for row in ws.iter_rows(min_row=2):
        print(row[0].value, row[1].value)  # indeksowanie od 0 (bo to zwykly tuple)
        # row[5].value = 1  # można edytować już istniejące komórki
        ws.cell(row=row[0].row, column=9).value =  row[5].value - row[7].value
    wb.save("C:\\Users\\kurs\\Documents\\temp\\dane1-mod.xlsx")  # problem: gubi format daty w pliku wyjsciowym
    wb.close()

if __name__ == '__main__':
    main()